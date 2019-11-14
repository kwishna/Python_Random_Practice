import openpyxl as xl

wb = xl.load_workbook('example.xlsx')
sheet1 = wb['Sheet1']
print(sheet1['A1'].value)
print(sheet1.title)
print(sheet1.cell(row=1, column=1).value)
print(sheet1.max_row, sheet1.max_column)
print()